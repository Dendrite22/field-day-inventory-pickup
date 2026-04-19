"""Rebuild First Pass.xlsx as a Plant & Equipment Contents inventory template."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("First Pass.xlsx")
ws = wb["Data"]

# ── Clear everything from row 1 down ─────────────────────────────────────────
# Unmerge all merged cells first
for merge in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(merge))
for row in ws.iter_rows():
    for cell in row:
        cell.value = None
        cell.font = Font()
        cell.fill = PatternFill()
        cell.alignment = Alignment()
        cell.border = Border()
        cell.number_format = "General"

# ── Column definitions ────────────────────────────────────────────────────────
COLS = [
    # (header, description, width)
    ("Unique ID",                   "System generated. Auto-increments from 100001 across entire portfolio.",         14),
    ("Asset ID (Client)",           "Imported from client records or input on inspection.",                           16),
    ("Parent Asset",                "Building / facility name. Imported or input on inspection. Autofill enabled.",   18),
    ("Address",                     "Imported from client list. Autofill enabled.",                                   22),
    ("Level",                       "e.g. Ground Floor, Level 1, Basement, External. Autofill enabled.",              16),
    ("Sub Location",                "Room or area name, e.g. Board Room, Reception, Open Plan. Autofill enabled.",    22),
    ("Asset Category (L1)",         "Selection: Furniture / Information Technology / Audio Visual / Appliances & White Goods / Office Equipment / Fitness & Recreation / Medical & First Aid / Vehicles & Motorised Plant / Other", 24),
    ("Asset Category (L2)",         "Sub-category. Selection driven by L1 choice.",                                   24),
    ("Asset Category (L3)",         "Item type. Selection driven by L2 choice.",                                      24),
    ("Asset Category (L4)",         "Variant or specification. Selection or '--' if not applicable.",                 18),
    ("Item Description",            "Free-text description of the specific item.",                                    28),
    ("Condition Rating (0-5)",      "Selection only: 0=Not Present, 1=As New, 2=Good, 3=Fair, 4=Poor, 5=Very Poor. Blank not accepted.", 14),
    ("Condition Rating Label",      "Auto-derived from Condition Rating value.",                                       16),
    ("Photo 1",                     "Camera capture — primary item photo.",                                           20),
    ("Photo 2",                     "Camera capture — secondary / detail photo.",                                     20),
    ("Quantity",                    "Count of identical items in this entry. Numerical only.",                         10),
    ("Unit of Measurement",         "Selection: ea / set / pair / lot.",                                              14),
    ("Width (mm)",                  "External width in millimetres. Required for furniture.",                          12),
    ("Depth (mm)",                  "External depth in millimetres. Required for furniture.",                          12),
    ("Height (mm)",                 "External height in millimetres.",                                                 12),
    ("Screen Size (inches)",        "Diagonal screen size in inches. Applies to TVs, monitors, projectors.",           14),
    ("Make / Brand",                "Manufacturer or brand name. Autofill enabled.",                                  18),
    ("Model",                       "Model name or number.",                                                          18),
    ("Serial Number",               "Manufacturer serial number. Critical for insurance schedules.",                   20),
    ("Asset Tag / Label",           "Client's own asset tag or label number.",                                        16),
    ("Colour / Finish",             "e.g. Black, White, Timber Veneer, Chrome, Grey Fabric. Autofill enabled.",       18),
    ("Comments / Recommendations",  "Free-text observations or recommendations.",                                     28),
    ("Defect (Yes/No)",             "Selection: Yes or No. Blank not accepted.",                                      14),
    ("Works Description",           "Active only when Defect = Yes.",                                                 28),
    ("Defect Priority",             "Selection: Immediate (High) / High / Moderate / Low. Active when Defect = Yes.", 20),
    ("Probable Repair Cost ($)",    "Estimated cost of defect rectification. Active when Defect = Yes.",              16),
    ("Works Photo 1",               "Camera capture — defect photo. Enabled when Defect = Yes.",                     20),
    ("Works Photo 2",               "Camera capture — works photo. Enabled when Defect = Yes.",                      20),
    ("Unit Rate - RCN ($)",          "Replacement Cost New per unit. From cost library or manual override on site.",   18),
    ("Estimated Replacement Cost ($)", "Formula: Unit Rate × Quantity.",                                             22),
    ("Remaining Useful Life (%)",   "Derived from Condition Rating: 1=100%, 2=75%, 3=50%, 4=25%, 5=0%, 0=N/A.",      18),
    ("Indemnity Value ($)",         "Formula: Estimated Replacement Cost × Remaining Useful Life. For financial reporting.", 18),
    ("Insurance Schedule Category","Selection: Contents / Plant & Equipment / Motor Vehicle / Other.",               24),
    ("Effective Life (years)",      "Total expected useful life for the item type.",                                   16),
]

NUM_COLS = len(COLS)

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL    = PatternFill("solid", fgColor="1F4E79")   # dark blue
DESC_FILL      = PatternFill("solid", fgColor="2E75B6")   # mid blue
SECTION_FILL   = PatternFill("solid", fgColor="D6E4F0")   # light blue tint
WHITE_FONT     = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
HEADER_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
DESC_FONT      = Font(name="Calibri", italic=True, color="FFFFFF", size=9)
BODY_FONT      = Font(name="Calibri", size=10)
BOLD_BODY      = Font(name="Calibri", bold=True, size=10)
WRAP           = Alignment(wrap_text=True, vertical="top")
CENTER         = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN           = Side(style="thin", color="BFBFBF")
THIN_BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COND_LABEL = {0: "Not Present", 1: "As New", 2: "Good", 3: "Fair", 4: "Poor", 5: "Very Poor"}
RUL_MAP    = {0: 0, 1: 100, 2: 75, 3: 50, 4: 25, 5: 0}

# ── Row 1 — App description ───────────────────────────────────────────────────
r1 = ws.cell(row=1, column=1)
r1.value = (
    "FIELD DAY — INVENTORY PICKUP | Plant & Equipment (Contents) Inspection Template. "
    "Select a building from the portfolio, then navigate the location hierarchy: "
    "Parent Asset → Address → Level → Sub Location → Asset Category (L1 to L4). "
    "Complete all fields in the item form before moving to the next entry. "
    "Autofill is active on all non-measurement, non-condition fields. "
    "Photos: Photo 1 & 2 (always available). Works Photos 1 & 2 unlock when Defect = Yes. "
    "GPS coordinates and map view (Google hybrid) are captured automatically per building."
)
r1.font = Font(name="Calibri", bold=True, color="1F4E79", size=11)
r1.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[1].height = 60
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)

# ── Row 2 — Autofill note ─────────────────────────────────────────────────────
r2 = ws.cell(row=2, column=1)
r2.value = (
    "AUTOFILL RULE: All non-measurement and non-condition fields offer autofill suggestions "
    "drawn from existing entries within the same building. Valuation fields (Unit Rate, ERC, "
    "Indemnity Value, RUL) are system-generated or manually overridden on site."
)
r2.font = Font(name="Calibri", italic=True, color="595959", size=9)
r2.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 36
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)

# ── Row 3 — blank separator ───────────────────────────────────────────────────
ws.row_dimensions[3].height = 6

# ── Row 4 — Column descriptions ──────────────────────────────────────────────
ws.row_dimensions[4].height = 72
for col_idx, (_, desc, _) in enumerate(COLS, start=1):
    c = ws.cell(row=4, column=col_idx, value=desc)
    c.font = DESC_FONT
    c.fill = DESC_FILL
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.border = THIN_BORDER

# ── Row 5 — Column headers ────────────────────────────────────────────────────
ws.row_dimensions[5].height = 36
for col_idx, (header, _, width) in enumerate(COLS, start=1):
    c = ws.cell(row=5, column=col_idx, value=header)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ── Data rows (starting row 6) ────────────────────────────────────────────────
# Each tuple:
# (uid, asset_id_client, parent_asset, address, level, sub_location,
#  l1, l2, l3, l4,
#  item_desc, cond_rating,
#  photo1, photo2,
#  qty, uom,
#  width_mm, depth_mm, height_mm, screen_in,
#  make, model, serial, asset_tag, colour,
#  comments,
#  defect, works_desc, defect_priority, probable_cost, works_photo1, works_photo2,
#  unit_rate,
#  insurance_cat, effective_life)

DATA = [
    # ── Reception ──────────────────────────────────────────────────────────────
    (100001,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Reception",
     "Furniture","Seating","Visitor Chair","--",
     "4-leg stackable visitor chair",2,
     None,None,4,"ea",520,560,820,None,"Haworth",None,None,None,"Charcoal Fabric",None,
     "No",None,None,None,None,None,420,"Contents",12),

    (100002,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Reception",
     "Furniture","Desks & Workstations","Reception Desk / Counter","--",
     "L-shape reception counter with modesty panel",2,
     None,None,1,"ea",2400,800,1100,None,"Custom Joinery",None,None,None,"White Laminate",None,
     "No",None,None,None,None,None,5500,"Contents",15),

    (100003,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Reception",
     "Audio Visual","Display Systems","Flat Screen TV","--",
     "Wall-mounted flat screen TV",2,
     None,None,1,"ea",None,None,None,65,"Samsung","QN65Q80C","SN-SAM-01",None,"Black",None,
     "No",None,None,None,None,None,2400,"Plant & Equipment",8),

    (100004,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Reception",
     "Furniture","Seating","Lounge Chair / Sofa","3-Seater",
     "3-seater lounge sofa",3,
     None,None,1,"ea",2100,900,800,None,None,None,None,None,"Grey Fabric",None,
     "No",None,None,None,None,None,2400,"Contents",12),

    (100005,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Reception",
     "Furniture","Storage & Filing","Storage Cabinet","--",
     "2-door storage cabinet with adjustable shelves",2,
     None,None,1,"ea",900,450,1800,None,None,None,None,None,"White","Reception stationery storage",
     "No",None,None,None,None,None,720,"Contents",20),

    # ── Main Administration Office ─────────────────────────────────────────────
    (100006,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Main Administration Office",
     "Furniture","Desks & Workstations","Straight Desk","--",
     "Straight office desk with cable management",2,
     None,None,3,"ea",1800,750,720,None,None,None,None,None,"Timber Veneer",None,
     "No",None,None,None,None,None,1050,"Contents",15),

    (100007,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Main Administration Office",
     "Furniture","Seating","Task Chair","--",
     "Ergonomic task chair with lumbar support and adjustable arms",2,
     None,None,3,"ea",670,640,920,None,"Buro","Metro II",None,None,"Black Mesh",None,
     "No",None,None,None,None,None,850,"Contents",12),

    (100008,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Main Administration Office",
     "Furniture","Storage & Filing","Filing Cabinet (Lateral)","4-Drawer",
     "4-drawer lateral filing cabinet with lock",3,
     None,None,2,"ea",900,580,1330,None,"Brownbuilt",None,None,None,"Beige",None,
     "No",None,None,None,None,None,950,"Contents",20),

    (100009,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Main Administration Office",
     "Furniture","Storage & Filing","Filing Cabinet (Pedestal)","3-Drawer",
     "Mobile 3-drawer pedestal with castors and lock",2,
     None,None,3,"ea",460,570,690,None,None,None,None,None,"White",None,
     "No",None,None,None,None,None,520,"Contents",20),

    (100010,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Main Administration Office",
     "Information Technology","Computing","Desktop Computer (Tower)","--",
     "Desktop PC tower",2,
     None,None,3,"ea",None,None,400,None,"Dell","OptiPlex 7010",None,None,"Black",None,
     "No",None,None,None,None,None,1850,"Plant & Equipment",5),

    (100011,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Main Administration Office",
     "Information Technology","Display & Peripherals","Monitor / Screen","--",
     "27-inch IPS LED monitor",2,
     None,None,3,"ea",None,None,None,27,"Dell","P2722H",None,None,"Black",None,
     "No",None,None,None,None,None,580,"Plant & Equipment",5),

    (100012,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Main Administration Office",
     "Information Technology","Display & Peripherals","Printer (Multifunction)","--",
     "A3 colour multifunction printer/copier/scanner",3,
     None,None,1,"ea",580,560,680,None,"Canon","iR-ADV C3530i","SN-CAN-01",None,"White/Grey",None,
     "No",None,None,None,None,None,1600,"Plant & Equipment",8),

    (100013,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Main Administration Office",
     "Furniture","Seating","Visitor Chair","--",
     "Visitor chair with padded seat",2,
     None,None,2,"ea",520,560,820,None,None,None,None,None,"Charcoal Fabric",None,
     "No",None,None,None,None,None,420,"Contents",12),

    # ── Board Room ─────────────────────────────────────────────────────────────
    (100014,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Board Room",
     "Furniture","Tables","Boardroom Table","--",
     "12-person boardroom table, boat-shaped",2,
     None,None,1,"ea",3600,1200,750,None,"Custom Joinery",None,None,None,"Dark Timber Veneer",None,
     "No",None,None,None,None,None,9500,"Contents",20),

    (100015,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Board Room",
     "Furniture","Seating","Meeting Chair","--",
     "High-back meeting chair with chrome base",2,
     None,None,12,"ea",550,590,1150,None,"Haworth","Fern",None,None,"Black Mesh",None,
     "No",None,None,None,None,None,520,"Contents",12),

    (100016,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Board Room",
     "Audio Visual","Display Systems","Flat Screen TV","--",
     "Wall-mounted 75-inch UHD display panel",2,
     None,None,2,"ea",None,None,None,75,"LG","75QNED90","SN-LG-01",None,"Black",None,
     "No",None,None,None,None,None,3400,"Plant & Equipment",8),

    (100017,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Board Room",
     "Audio Visual","Video Conferencing","Video Bar (All-in-One)","--",
     "Integrated video conferencing bar with camera, mic and speaker",2,
     None,None,1,"ea",700,150,100,None,"Poly","Studio E70","SN-POLY-01",None,"Black",None,
     "No",None,None,None,None,None,9200,"Plant & Equipment",8),

    (100018,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Board Room",
     "Furniture","Storage & Filing","Credenza / Buffet","--",
     "2-door credenza with timber veneer finish",2,
     None,None,2,"ea",1800,450,750,None,None,None,None,None,"Dark Timber Veneer",None,
     "No",None,None,None,None,None,1600,"Contents",20),

    # ── Kitchen ────────────────────────────────────────────────────────────────
    (100019,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Kitchen",
     "Appliances & White Goods","Kitchen Appliances","Refrigerator / Fridge","--",
     "French door refrigerator with water dispenser",2,
     None,None,1,"ea",750,680,1780,None,"Fisher & Paykel","RF540ADUB5","SN-FP-01",None,"Stainless Steel",None,
     "No",None,None,None,None,None,1900,"Plant & Equipment",12),

    (100020,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Kitchen",
     "Appliances & White Goods","Kitchen Appliances","Dishwasher","--",
     "Under-bench dishwasher, freestanding",3,
     None,None,1,"ea",600,610,850,None,"Bosch","SMS68TI01A","SN-BSH-01",None,"White",None,
     "No",None,None,None,None,None,1450,"Plant & Equipment",12),

    (100021,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Kitchen",
     "Appliances & White Goods","Kitchen Appliances","Microwave Oven","--",
     "Countertop microwave oven",2,
     None,None,1,"ea",510,430,310,None,"Panasonic","NN-ST664W",None,None,"White",None,
     "No",None,None,None,None,None,420,"Plant & Equipment",10),

    (100022,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Kitchen",
     "Appliances & White Goods","Kitchen Appliances","Coffee Machine (Auto)","--",
     "Fully automatic bean-to-cup coffee machine",2,
     None,None,1,"ea",280,380,360,None,"Jura","E8","SN-JURA-01",None,"Silver/Black","Includes integrated milk frother",
     "No",None,None,None,None,None,1600,"Plant & Equipment",10),

    (100023,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Kitchen",
     "Appliances & White Goods","Kitchen Appliances","Water Cooler / Dispenser","--",
     "Plumbed-in instant hot/cold/sparkling water tap",2,
     None,None,1,"ea",300,370,1100,None,"Zip","HydroTap G5","SN-ZIP-01",None,"Chrome/Black",None,
     "No",None,None,None,None,None,950,"Plant & Equipment",12),

    (100024,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Kitchen",
     "Furniture","Tables","Dining Table","--",
     "Rectangular dining/breakout table",3,
     None,None,1,"ea",1800,900,750,None,None,None,None,None,"Timber Veneer",None,
     "No",None,None,None,None,None,1200,"Contents",20),

    (100025,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Kitchen",
     "Furniture","Seating","Bar Stool","--",
     "Counter-height bar stool with footrest",2,
     None,None,4,"ea",420,400,770,None,None,None,None,None,"Charcoal Fabric",None,
     "No",None,None,None,None,None,380,"Contents",12),

    # ── Open Plan ──────────────────────────────────────────────────────────────
    (100026,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Open Plan",
     "Furniture","Desks & Workstations","Height-Adjustable Desk","Electric",
     "Electric sit-stand desk with digital height memory",2,
     None,None,6,"ea",1600,800,650,None,"Ergotron","WorkFit-D",None,None,"White","Height range 650-1280mm",
     "No",None,None,None,None,None,2400,"Contents",15),

    (100027,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Open Plan",
     "Furniture","Seating","Task Chair","--",
     "High-performance ergonomic task chair",2,
     None,None,6,"ea",680,650,920,None,"Herman Miller","Aeron",None,None,"Black Mesh",None,
     "No",None,None,None,None,None,1850,"Contents",12),

    (100028,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Open Plan",
     "Information Technology","Display & Peripherals","Monitor / Screen","--",
     "27-inch USB-C IPS monitor",2,
     None,None,12,"ea",None,None,None,27,"Dell","U2722D",None,None,"Black",None,
     "No",None,None,None,None,None,580,"Plant & Equipment",5),

    (100029,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Open Plan",
     "Furniture","Screens & Partitions","Acoustic Screen","--",
     "Freestanding acoustic desk divider screen",2,
     None,None,8,"ea",1200,50,1200,None,None,None,None,None,"Grey Fabric",None,
     "No",None,None,None,None,None,480,"Contents",15),

    (100030,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Open Plan",
     "Furniture","Seating","Visitor Chair","--",
     "4-leg visitor chair with upholstered seat",3,
     None,None,4,"ea",520,560,820,None,None,None,None,None,"Charcoal Fabric",None,
     "No",None,None,None,None,None,420,"Contents",12),

    (100031,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Open Plan",
     "Information Technology","Computing","Laptop / Notebook","--",
     "Business laptop with docking station",2,
     None,None,6,"ea",320,230,18,None,"Dell","Latitude 5540",None,None,"Black",None,
     "No",None,None,None,None,None,2100,"Plant & Equipment",5),

    # ── Store Room ─────────────────────────────────────────────────────────────
    (100032,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Store Room",
     "Furniture","Storage & Filing","Bookcase / Shelf Unit","--",
     "5-shelf steel shelf unit, heavy duty",3,
     None,None,6,"ea",900,380,1980,None,None,None,None,None,"White",None,
     "No",None,None,None,None,None,620,"Contents",20),

    (100033,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Store Room",
     "Furniture","Storage & Filing","Locker","Quad",
     "4-door staff locker with padlock provision",4,
     None,None,3,"ea",900,450,1800,None,"Brownbuilt",None,None,None,"Grey","Doors showing rust and impact damage — repaint or replace",
     "Yes","Doors rusted and misaligned — replace locker bays","Moderate",1950,None,None,650,"Contents",20),

    # ── Lobby ──────────────────────────────────────────────────────────────────
    (100034,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Lobby",
     "Furniture","Seating","Lounge Chair / Sofa","2-Seater",
     "2-seater lounge sofa",2,
     None,None,2,"ea",1600,850,780,None,None,None,None,None,"Navy Fabric",None,
     "No",None,None,None,None,None,1900,"Contents",12),

    (100035,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Lobby",
     "Furniture","Seating","Lounge Chair / Sofa","1-Seater",
     "Single-seat armchair",2,
     None,None,2,"ea",900,850,780,None,None,None,None,None,"Navy Fabric",None,
     "No",None,None,None,None,None,1100,"Contents",12),

    (100036,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Lobby",
     "Audio Visual","Display Systems","Digital Signage Screen","--",
     "Commercial-grade digital signage display, landscape orientation",2,
     None,None,1,"ea",None,None,None,55,"Samsung","QH55B","SN-SAM-02",None,"Black",None,
     "No",None,None,None,None,None,2800,"Plant & Equipment",8),

    (100037,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Lobby",
     "Furniture","Soft Furnishings","Artwork / Print","--",
     "Framed architectural print, professionally mounted",2,
     None,None,4,"ea",600,50,800,None,None,None,None,None,"Black Frame",None,
     "No",None,None,None,None,None,350,"Contents",20),

    # ── Open Plan — defect example ─────────────────────────────────────────────
    (100038,"BO214","Clubrooms","50 Shephard Street","Ground Floor","Open Plan",
     "Office Equipment","Mail & Document","Shredder","--",
     "Cross-cut paper shredder, P-4 security level",4,
     None,None,1,"ea",360,230,580,None,"Fellowes","Powershred 79Ci","SN-FEL-01",None,"Black/Silver",None,
     "Yes","Jammed mechanism — unit fails to power on. Replace.","Moderate",520,None,None,520,"Plant & Equipment",10),
]

# ── Write data rows ───────────────────────────────────────────────────────────
DATA_START = 6
for row_offset, row_data in enumerate(DATA):
    r = DATA_START + row_offset
    ws.row_dimensions[r].height = 18
    (uid, asset_id, parent, address, level, sub_loc,
     l1, l2, l3, l4,
     item_desc, cond,
     photo1, photo2,
     qty, uom,
     w_mm, d_mm, h_mm, screen,
     make, model, serial, tag, colour,
     comments,
     defect, works_desc, defect_pri, probable_cost, works_p1, works_p2,
     unit_rate,
     ins_cat, eff_life) = row_data

    cond_label = COND_LABEL.get(cond, "")
    rul_pct    = RUL_MAP.get(cond, 0) if cond != 0 else None
    erc_col    = get_column_letter(35)   # Estimated Replacement Cost column
    rul_col    = get_column_letter(36)   # RUL column

    vals = [
        uid, asset_id, parent, address, level, sub_loc,
        l1, l2, l3, l4,
        item_desc, cond, cond_label,
        photo1, photo2,
        qty, uom,
        w_mm, d_mm, h_mm, screen,
        make, model, serial, tag, colour,
        comments,
        defect, works_desc, defect_pri, probable_cost, works_p1, works_p2,
        unit_rate,
        f"={get_column_letter(34)}{r}*{get_column_letter(16)}{r}",   # ERC = unit_rate * qty
        rul_pct,
        f"={erc_col}{r}*{rul_col}{r}/100" if rul_pct is not None else "N/A",  # Indemnity
        ins_cat, eff_life,
    ]

    for col_idx, val in enumerate(vals, start=1):
        c = ws.cell(row=r, column=col_idx, value=val)
        c.font = BODY_FONT
        c.border = THIN_BORDER
        c.alignment = Alignment(vertical="center", wrap_text=False)

        # Number formatting
        col_letter = get_column_letter(col_idx)
        header = COLS[col_idx - 1][0]
        if "($)" in header or "Rate" in header or "Cost" in header or "Value" in header:
            c.number_format = '"$"#,##0'
        elif "(%)" in header:
            c.number_format = "0%"
            if isinstance(val, (int, float)):
                c.value = val / 100   # store as decimal for % format

    # Alternating row fill
    if row_offset % 2 == 0:
        fill = PatternFill("solid", fgColor="F2F7FC")
        for col_idx in range(1, NUM_COLS + 1):
            ws.cell(row=r, column=col_idx).fill = fill

# ── Freeze panes below header ─────────────────────────────────────────────────
ws.freeze_panes = "A6"

# ── Auto-filter on header row ─────────────────────────────────────────────────
ws.auto_filter.ref = f"A5:{get_column_letter(NUM_COLS)}5"

# ── Save ───────────────────────────────────────────────────────────────────────
wb.save("First Pass.xlsx")
print(f"Done. Written {len(DATA)} data rows across {NUM_COLS} columns.")
